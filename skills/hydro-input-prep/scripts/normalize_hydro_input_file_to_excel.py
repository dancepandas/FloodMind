#!/usr/bin/env python3
"""把原始输入文件自动清洗为标准中间 Excel。"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[3]
SCRIPT_DIR = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from openpyxl import load_workbook

from _common import (
    DEFAULT_STATION_CODE,
    expand_rows_by_station,
    infer_durations,
    infer_forecast_time,
    normalize_rainfall_rows,
    parse_model_types,
    read_standard_workbook,
    write_standard_workbook,
)
from aojiang_station_resolver import describe_station_codes, resolve_aojiang_station_codes, resolve_aojiang_station_name

TIME_HEADERS = ["time", "datetime", "时间", "时刻", "date_time", "时间点", "日期时间"]
RAINFALL_HEADERS = ["rainfallValue", "rainfall", "降雨", "雨量", "precipitation", "计算面降雨量", "面雨量", "降雨量"]
STATION_HEADERS = ["stationCode", "station_code", "站点编码", "站号", "station"]
STATION_NAME_HEADERS = ["stationName", "sectionName", "station_name", "section_name", "站点名称", "断面名称", "站名", "断面"]
PHASE_HEADERS = ["phase", "type", "period", "阶段", "数据类型"]


def find_header(row: list[str], candidates: list[str]) -> str | None:
    lowered = {str(cell).strip().lower(): str(cell).strip() for cell in row if str(cell).strip()}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    return None


def detect_table_rows(raw_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    if not raw_rows:
        raise ValueError("输入文件中没有可用数据")
    sample_keys = list(raw_rows[0].keys())
    time_key = find_header(sample_keys, TIME_HEADERS)
    rainfall_key = find_header(sample_keys, RAINFALL_HEADERS)
    station_key = find_header(sample_keys, STATION_HEADERS)
    station_name_key = find_header(sample_keys, STATION_NAME_HEADERS)
    phase_key = find_header(sample_keys, PHASE_HEADERS)

    if not time_key or not rainfall_key:
        raise ValueError("未识别到时间列或降雨列，请确保原始文件包含 time/rainfallValue 语义列")

    notes = [
        f"识别时间列: {time_key}",
        f"识别降雨列: {rainfall_key}",
        f"识别站点列: {station_key or '未识别，后续将使用默认 stationCode'}",
        f"识别站点名称列: {station_name_key or '未识别'}",
        f"识别阶段列: {phase_key or '未识别'}",
    ]

    normalized: list[dict[str, Any]] = []
    for row in raw_rows:
        normalized.append(
            {
                "time": row.get(time_key),
                "rainfallValue": row.get(rainfall_key),
                "stationCode": row.get(station_key) if station_key else None,
                "stationName": row.get(station_name_key) if station_name_key else None,
                "phase": row.get(phase_key) if phase_key else None,
            }
        )
    return normalized, notes


def _resolve_explicit_header(sample_keys: list[str], explicit_name: str | None) -> str | None:
    if explicit_name in (None, ""):
        return None
    explicit = str(explicit_name).strip()
    for key in sample_keys:
        if str(key).strip() == explicit:
            return str(key).strip()
    raise ValueError(f"指定列名不存在: {explicit}")


def detect_table_rows_with_mapping(
    raw_rows: list[dict[str, Any]],
    *,
    time_column: str | None = None,
    rainfall_column: str | None = None,
    station_column: str | None = None,
    station_name_column: str | None = None,
    phase_column: str | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    if not raw_rows:
        raise ValueError("输入文件中没有可用数据")

    sample_keys = list(raw_rows[0].keys())
    explicit_time_key = _resolve_explicit_header(sample_keys, time_column)
    explicit_rainfall_key = _resolve_explicit_header(sample_keys, rainfall_column)
    explicit_station_key = _resolve_explicit_header(sample_keys, station_column)
    explicit_station_name_key = _resolve_explicit_header(sample_keys, station_name_column)
    explicit_phase_key = _resolve_explicit_header(sample_keys, phase_column)

    if any(value is not None for value in (explicit_time_key, explicit_rainfall_key, explicit_station_key, explicit_station_name_key, explicit_phase_key)):
        time_key = explicit_time_key or find_header(sample_keys, TIME_HEADERS)
        rainfall_key = explicit_rainfall_key or find_header(sample_keys, RAINFALL_HEADERS)
        station_key = explicit_station_key or find_header(sample_keys, STATION_HEADERS)
        station_name_key = explicit_station_name_key or find_header(sample_keys, STATION_NAME_HEADERS)
        phase_key = explicit_phase_key or find_header(sample_keys, PHASE_HEADERS)
        if not time_key or not rainfall_key:
            raise ValueError("指定列映射后，仍未识别到时间列或降雨列")

        notes = [
            f"显式指定时间列: {time_key}" if explicit_time_key else f"识别时间列: {time_key}",
            f"显式指定降雨列: {rainfall_key}" if explicit_rainfall_key else f"识别降雨列: {rainfall_key}",
            f"显式指定站点列: {station_key}" if explicit_station_key else f"识别站点列: {station_key or '未识别，后续将使用默认 stationCode'}",
            f"显式指定站点名称列: {station_name_key}" if explicit_station_name_key else f"识别站点名称列: {station_name_key or '未识别'}",
            f"显式指定阶段列: {phase_key}" if explicit_phase_key else f"识别阶段列: {phase_key or '未识别'}",
        ]
        normalized: list[dict[str, Any]] = []
        for row in raw_rows:
            normalized.append(
                {
                    "time": row.get(time_key),
                    "rainfallValue": row.get(rainfall_key),
                    "stationCode": row.get(station_key) if station_key else None,
                    "stationName": row.get(station_name_key) if station_name_key else None,
                    "phase": row.get(phase_key) if phase_key else None,
                }
            )
        return normalized, notes

    return detect_table_rows(raw_rows)


def apply_aojiang_station_semantics(
    rows: list[dict[str, Any]],
    *,
    case_name: str,
    default_station_code: str,
    task_description: str | None,
) -> tuple[list[dict[str, Any]], list[str]]:
    notes: list[str] = []
    if case_name.strip().lower() != "aojiang":
        return rows, notes

    mapped_rows: list[dict[str, Any]] = []
    mapped_count = 0
    unresolved_names: list[str] = []
    has_station_code = any(str(row.get("stationCode") or "").strip() for row in rows)

    for row in rows:
        station_code = str(row.get("stationCode") or "").strip()
        if station_code:
            mapped_rows.append(row)
            continue
        station_name = str(row.get("stationName") or "").strip()
        if station_name:
            resolved = resolve_aojiang_station_name(station_name)
            if resolved:
                mapped_rows.append({**row, "stationCode": resolved})
                mapped_count += 1
            else:
                unresolved_names.append(station_name)
                mapped_rows.append(row)
            continue
        mapped_rows.append(row)

    if mapped_count:
        notes.append(f"已根据敖江站点名称列自动映射 {mapped_count} 行 stationCode。")
    if unresolved_names:
        unique_names = []
        for name in unresolved_names:
            if name not in unique_names:
                unique_names.append(name)
        notes.append(f"以下站点名称未能映射到敖江 stationCode: {', '.join(unique_names)}")

    has_station_code = has_station_code or any(str(row.get("stationCode") or "").strip() for row in mapped_rows)
    if has_station_code:
        return mapped_rows, notes

    if not task_description:
        notes.append(f"未识别到站点列，回退默认 stationCode={default_station_code}。")
        return mapped_rows, notes

    station_codes, resolve_notes = resolve_aojiang_station_codes(task_description)
    expanded_rows: list[dict[str, Any]] = []
    for row in mapped_rows:
        for station_code in station_codes:
            expanded_rows.append({**row, "stationCode": station_code})
    notes.append(f"未识别到站点列，已根据任务描述展开 stationCode: {describe_station_codes(station_codes)}")
    notes.extend(resolve_notes)
    return expanded_rows, notes


def split_rows(
    rows: list[dict[str, Any]],
    *,
    forecast_time: str | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    explicit_history: list[dict[str, Any]] = []
    explicit_forecast: list[dict[str, Any]] = []
    notes: list[str] = []
    has_phase = any(row.get("phase") not in (None, "") for row in rows)

    if has_phase:
        for row in rows:
            phase = str(row.get("phase") or "").strip().lower()
            if phase in {"history", "historical", "history_rainfall", "历史"}:
                explicit_history.append(row)
            elif phase in {"forecast", "future", "forecast_rainfall", "预报", "未来"}:
                explicit_forecast.append(row)
        if explicit_history or explicit_forecast:
            notes.append("按阶段列拆分为历史降雨和未来降雨")
            return explicit_history, explicit_forecast, notes

    if not forecast_time:
        raise ValueError("原始文件未提供阶段列时，必须显式传入 --forecast_time 用于拆分历史/未来数据")

    for row in rows:
        time_value = str(row.get("time") or "").strip()
        if time_value <= forecast_time:
            explicit_history.append(row)
        else:
            explicit_forecast.append(row)
    notes.append("按 forecast_time 拆分历史降雨和未来降雨")
    return explicit_history, explicit_forecast, notes


def load_csv_like(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;") if sample else csv.excel
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        return [dict(row) for row in reader]


def load_json_rows(path: Path) -> tuple[dict[str, Any] | None, list[dict[str, Any]] | None]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and {"forecastTime", "historyDuration", "futureDuration", "modelDataParams", "modelForecastRainfallParams"}.issubset(payload.keys()):
        return payload, None
    if isinstance(payload, list) and all(isinstance(item, dict) for item in payload):
        return None, payload
    if isinstance(payload, dict) and "rows" in payload and isinstance(payload["rows"], list):
        rows = payload["rows"]
        if all(isinstance(item, dict) for item in rows):
            return None, rows
    raise ValueError("JSON 文件既不是标准 input.json，也不是可识别的对象数组")


def load_excel_rows(path: Path) -> tuple[dict[str, Any] | None, list[dict[str, Any]] | None]:
    workbook = load_workbook(path, data_only=True)
    if {"metadata", "history_rainfall", "forecast_rainfall"}.issubset(workbook.sheetnames):
        return {"standard_workbook": True}, None

    sheet = workbook[workbook.sheetnames[0]]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        raise ValueError("Excel 文件为空")
    headers = ["" if item is None else str(item).strip() for item in values[0]]
    rows: list[dict[str, Any]] = []
    for raw_row in values[1:]:
        if raw_row is None:
            continue
        row = {headers[index]: raw_row[index] for index in range(min(len(headers), len(raw_row))) if headers[index]}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return None, rows


def main() -> None:
    parser = argparse.ArgumentParser(description="把原始输入文件自动清洗为标准中间 Excel")
    parser.add_argument("--input_file", required=True, help="原始文件路径，支持 xlsx/csv/json/txt")
    parser.add_argument("--output_file", required=True, help="输出标准中间 Excel 路径")
    parser.add_argument("--case_name", default="aojiang", help="案例名称")
    parser.add_argument("--forecast_time", help="用于拆分历史/未来的锚点时间")
    parser.add_argument("--station_code", default=DEFAULT_STATION_CODE, help="默认 stationCode")
    parser.add_argument("--model_types", default="XAJ", help="模型类型，逗号分隔")
    parser.add_argument("--source_note", default="", help="附加备注")
    parser.add_argument("--task_description", help="任务描述；当原始文件无 stationCode 时，可用于按业务语义展开站点")
    parser.add_argument("--time_column", help="显式指定时间列名")
    parser.add_argument("--rainfall_column", help="显式指定降雨列名")
    parser.add_argument("--station_column", help="显式指定 stationCode 列名")
    parser.add_argument("--station_name_column", help="显式指定站点名称列名")
    parser.add_argument("--phase_column", help="显式指定阶段列名")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    suffix = input_path.suffix.lower()
    notes: list[str] = []
    model_types = parse_model_types(args.model_types)

    if suffix in {".xlsx", ".xlsm"}:
        metadata_or_marker, rows = load_excel_rows(input_path)
        if metadata_or_marker and metadata_or_marker.get("standard_workbook"):
            workbook_data = read_standard_workbook(input_path)
            metadata = workbook_data["metadata"]
            metadata["caseName"] = metadata.get("caseName") or args.case_name
            metadata["defaultStationCode"] = metadata.get("defaultStationCode") or args.station_code
            metadata["modelTypes"] = metadata.get("modelTypes") or ",".join(model_types)
            if args.source_note:
                workbook_data["notes"].append(args.source_note)
            write_standard_workbook(
                Path(args.output_file),
                metadata=metadata,
                history_rows=workbook_data["history_rows"],
                forecast_rows=workbook_data["forecast_rows"],
                notes=workbook_data["notes"],
            )
            print(f"标准中间 Excel 已创建: {args.output_file}")
            print("source type: standard_workbook")
            return
    elif suffix in {".csv", ".txt", ".tsv"}:
        rows = load_csv_like(input_path)
    elif suffix == ".json":
        payload, rows = load_json_rows(input_path)
        if payload is not None:
            metadata = {
                "caseName": args.case_name,
                "forecastTime": payload["forecastTime"],
                "historyDuration": payload["historyDuration"],
                "futureDuration": payload["futureDuration"],
                "timeStepHours": 1,
                "defaultStationCode": args.station_code,
                "modelTypes": ",".join(payload.get("modelRunParam", {}).keys()) or ",".join(model_types),
                "sourceType": "input_json",
                "sourceNote": args.source_note or "由现有 input.json 反向整理",
            }
            history_rows = normalize_rainfall_rows(payload.get("modelDataParams", []), args.station_code)
            forecast_rows = normalize_rainfall_rows(payload.get("modelForecastRainfallParams", []), args.station_code)
            notes = ["输入文件已识别为标准 input.json，并已转换为中间 Excel 便于检查。"]
            if args.source_note:
                notes.append(args.source_note)
            write_standard_workbook(Path(args.output_file), metadata=metadata, history_rows=history_rows, forecast_rows=forecast_rows, notes=notes)
            print(f"标准中间 Excel 已创建: {args.output_file}")
            print("source type: input_json")
            return
    else:
        raise SystemExit(f"暂不支持的文件类型: {suffix}")

    if rows is None:
        raise SystemExit("未读取到可处理的数据行")

    normalized_rows, detect_notes = detect_table_rows_with_mapping(
        rows,
        time_column=args.time_column,
        rainfall_column=args.rainfall_column,
        station_column=args.station_column,
        station_name_column=args.station_name_column,
        phase_column=args.phase_column,
    )
    normalized_rows, station_notes = apply_aojiang_station_semantics(
        normalized_rows,
        case_name=args.case_name,
        default_station_code=args.station_code,
        task_description=args.task_description,
    )
    history_rows, forecast_rows, split_notes = split_rows(normalized_rows, forecast_time=args.forecast_time)
    history_rows = normalize_rainfall_rows(history_rows, args.station_code)
    forecast_rows = normalize_rainfall_rows(forecast_rows, args.station_code)
    forecast_time = args.forecast_time or infer_forecast_time(history_rows, forecast_rows, 1)
    history_duration, future_duration = infer_durations(history_rows, forecast_rows, forecast_time, 1)
    history_rows, forecast_rows, fill_notes = expand_rows_by_station(
        history_rows=history_rows,
        forecast_rows=forecast_rows,
        forecast_time=forecast_time,
        history_duration=history_duration,
        future_duration=future_duration,
        time_step_hours=1,
        default_station_code=args.station_code,
    )

    metadata = {
        "caseName": args.case_name,
        "forecastTime": forecast_time,
        "historyDuration": history_duration,
        "futureDuration": future_duration,
        "timeStepHours": 1,
        "defaultStationCode": args.station_code,
        "modelTypes": ",".join(model_types),
        "sourceType": "raw_file",
        "sourceNote": args.source_note or f"由原始文件自动识别: {input_path.name}",
    }
    notes.extend(detect_notes)
    notes.extend(station_notes)
    notes.extend(split_notes)
    notes.extend(fill_notes)
    if args.source_note:
        notes.append(args.source_note)
    write_standard_workbook(Path(args.output_file), metadata=metadata, history_rows=history_rows, forecast_rows=forecast_rows, notes=notes)

    print(f"标准中间 Excel 已创建: {args.output_file}")
    print(f"history rows: {len(history_rows)}")
    print(f"forecast rows: {len(forecast_rows)}")
    print(f"models: {', '.join(model_types)}")


if __name__ == "__main__":
    main()
