#!/usr/bin/env python3
"""共享的水文输入中间 Excel 读写、补零与校验工具。"""

from __future__ import annotations

import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DEFAULT_STATION_CODE = "33c76b8bd9384486a945c2fc7fd622eb"
DEFAULT_MODEL_TYPES = ["XAJ"]
SUPPORTED_MODELS = {"XAJ", "GR4J", "HYMOD", "MASKGEN"}
DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
TIME_ALIGNMENT_TOLERANCE_MINUTES = 10
RESERVED_SHEETS = {"metadata", "notes", "history_rainfall", "forecast_rainfall"}
STATION_SHEET_HEADERS = ["phase", "time", "rainfallValue", "stationCode", "isAutoFilled"]

METADATA_KEYS = [
    "caseName",
    "forecastTime",
    "historyDuration",
    "futureDuration",
    "timeStepHours",
    "defaultStationCode",
    "stationCodes",
    "modelTypes",
    "sourceType",
    "sourceNote",
]


def parse_datetime(raw: Any) -> datetime:
    if isinstance(raw, datetime):
        return raw.replace(microsecond=0)
    try:
        return datetime.strptime(str(raw).strip(), DATETIME_FORMAT)
    except ValueError as exc:
        raise ValueError(f"时间格式错误，应为 {DATETIME_FORMAT}: {raw}") from exc


def format_datetime(value: datetime) -> str:
    return value.strftime(DATETIME_FORMAT)


def align_time_to_grid(raw_time: Any, anchor_time: str, time_step_hours: int, tolerance_minutes: int = TIME_ALIGNMENT_TOLERANCE_MINUTES) -> str:
    """把接近标准时点的时间吸附到时间网格上。"""
    dt = parse_datetime(raw_time)
    anchor = parse_datetime(anchor_time)
    step_seconds = int(timedelta(hours=time_step_hours).total_seconds())
    delta_seconds = (dt - anchor).total_seconds()
    nearest_step = round(delta_seconds / step_seconds)
    aligned = anchor + timedelta(seconds=nearest_step * step_seconds)
    tolerance_seconds = tolerance_minutes * 60
    if abs((dt - aligned).total_seconds()) <= tolerance_seconds:
        return format_datetime(aligned)
    return format_datetime(dt)


def align_rows_to_grid(
    rows: list[dict[str, Any]],
    *,
    forecast_time: str,
    time_step_hours: int,
    tolerance_minutes: int = TIME_ALIGNMENT_TOLERANCE_MINUTES,
) -> list[dict[str, Any]]:
    aligned_rows: list[dict[str, Any]] = []
    for row in rows:
        aligned_rows.append(
            {
                **row,
                "time": align_time_to_grid(row.get("time"), forecast_time, time_step_hours, tolerance_minutes),
            }
        )
    return aligned_rows


def parse_model_types(raw: str | list[str] | None) -> list[str]:
    if raw is None:
        return list(DEFAULT_MODEL_TYPES)
    if isinstance(raw, list):
        parts = [str(item).strip().upper() for item in raw if str(item).strip()]
    else:
        parts = [item.strip().upper() for item in str(raw).replace("，", ",").split(",") if item.strip()]
    if not parts:
        return list(DEFAULT_MODEL_TYPES)

    invalid = [item for item in parts if item not in SUPPORTED_MODELS]
    if invalid:
        raise ValueError(f"不支持的模型类型: {', '.join(invalid)}")
    return parts


def build_model_run_param(model_types: list[str]) -> dict[str, list[Any]]:
    return {model_type: [] for model_type in model_types}


def ensure_positive_int(raw: Any, field_name: str) -> int:
    try:
        value = int(raw)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} 必须是整数") from exc
    if value <= 0:
        raise ValueError(f"{field_name} 必须大于 0")
    return value


def normalize_station_code(value: Any, default_station_code: str) -> str:
    station_code = str(value).strip() if value not in (None, "") else default_station_code.strip()
    if not station_code:
        raise ValueError("stationCode 为空，且 metadata.defaultStationCode 也为空")
    return station_code


def coerce_rainfall_value(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"rainfallValue 不是合法数字: {value}") from exc


def normalize_rainfall_rows(rows: list[dict[str, Any]], default_station_code: str) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        raw_time = row.get("time")
        if raw_time in (None, ""):
            raise ValueError(f"第 {index} 行缺少 time")
        normalized.append(
            {
                "time": format_datetime(parse_datetime(raw_time)),
                "rainfallValue": coerce_rainfall_value(row.get("rainfallValue")),
                "stationCode": normalize_station_code(row.get("stationCode"), default_station_code),
                "phase": str(row.get("phase") or "").strip().lower(),
                "isAutoFilled": str(row.get("isAutoFilled") or "").strip(),
            }
        )
    return normalized


def autosize_sheet(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)


def group_rows_by_station(rows: list[dict[str, Any]], default_station_code: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        station_code = normalize_station_code(row.get("stationCode"), default_station_code)
        grouped.setdefault(station_code, []).append({**row, "stationCode": station_code})
    return grouped


def station_sheet_name(station_code: str) -> str:
    cleaned = "".join("_" if ch in '[]:*?/\\' else ch for ch in station_code).strip()
    return (cleaned or "station")[0:31]


def _sort_station_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    phase_order = {"history": 0, "forecast": 1}
    return sorted(rows, key=lambda row: (phase_order.get(str(row.get("phase") or ""), 9), str(row.get("time") or "")))


def write_standard_workbook(
    output_path: Path,
    *,
    metadata: dict[str, Any],
    history_rows: list[dict[str, Any]],
    forecast_rows: list[dict[str, Any]],
    notes: list[str] | None = None,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=11)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    station_rows = group_rows_by_station(
        [{**row, "phase": "history"} for row in history_rows] + [{**row, "phase": "forecast"} for row in forecast_rows],
        str(metadata.get("defaultStationCode") or DEFAULT_STATION_CODE),
    )
    metadata = dict(metadata)
    metadata["stationCodes"] = ",".join(station_rows.keys())
    final_notes = list(notes or [])

    metadata_sheet = workbook.create_sheet("metadata")
    metadata_sheet.append(["key", "value"])
    for cell in metadata_sheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment
    for key in METADATA_KEYS:
        metadata_sheet.append([key, metadata.get(key, "")])
    for row in metadata_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap_alignment
    metadata_sheet.freeze_panes = "A2"
    autosize_sheet(metadata_sheet)

    for station_code, rows in station_rows.items():
        sheet_title = station_sheet_name(station_code)
        if sheet_title != station_code:
            final_notes.append(f"stationCode={station_code} 的工作表名称因 Excel 31 字符限制被截断为 {sheet_title}，实际编码以 stationCode 列为准。")
        sheet = workbook.create_sheet(sheet_title)
        sheet.append(STATION_SHEET_HEADERS)
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = wrap_alignment
        for row in _sort_station_rows(rows):
            sheet.append([
                row.get("phase", ""),
                row.get("time", ""),
                row.get("rainfallValue", ""),
                row.get("stationCode", station_code),
                row.get("isAutoFilled", ""),
            ])
        for data_row in sheet.iter_rows(min_row=2):
            for cell in data_row:
                cell.font = body_font
                cell.alignment = wrap_alignment
        sheet.freeze_panes = "A2"
        autosize_sheet(sheet)

    notes_sheet = workbook.create_sheet("notes")
    notes_sheet.append(["note"])
    notes_sheet["A1"].font = header_font
    notes_sheet["A1"].alignment = wrap_alignment
    for note in final_notes:
        notes_sheet.append([note])
    for row in notes_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap_alignment
    notes_sheet.freeze_panes = "A2"
    autosize_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _read_key_value_sheet(sheet) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = "" if row[0] is None else str(row[0]).strip()
        if not key:
            continue
        value = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        result[key] = value
    return result


def _read_legacy_rainfall_sheet(sheet) -> list[dict[str, Any]]:
    headers = ["" if cell.value is None else str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers) if header}
    required = {"time", "rainfallValue", "stationCode"}
    if not required.issubset(header_index):
        raise ValueError(f"工作表 {sheet.title} 缺少必填列: {', '.join(sorted(required - set(header_index)))}")

    rows: list[dict[str, Any]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        if raw_row is None:
            continue
        time_value = raw_row[header_index["time"]] if len(raw_row) > header_index["time"] else None
        rainfall_value = raw_row[header_index["rainfallValue"]] if len(raw_row) > header_index["rainfallValue"] else None
        station_code = raw_row[header_index["stationCode"]] if len(raw_row) > header_index["stationCode"] else None
        if time_value in (None, "") and rainfall_value in (None, "") and station_code in (None, ""):
            continue
        rows.append({"time": time_value, "rainfallValue": rainfall_value, "stationCode": station_code, "isAutoFilled": ""})
    return rows


def _read_station_sheet(sheet, default_station_code: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    headers = ["" if cell.value is None else str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: idx for idx, header in enumerate(headers) if header}
    required = {"phase", "time", "rainfallValue"}
    if not required.issubset(header_index):
        raise ValueError(f"工作表 {sheet.title} 缺少必填列: {', '.join(sorted(required - set(header_index)))}")

    history_rows: list[dict[str, Any]] = []
    forecast_rows: list[dict[str, Any]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        if raw_row is None:
            continue
        phase = raw_row[header_index["phase"]] if len(raw_row) > header_index["phase"] else None
        time_value = raw_row[header_index["time"]] if len(raw_row) > header_index["time"] else None
        rainfall_value = raw_row[header_index["rainfallValue"]] if len(raw_row) > header_index["rainfallValue"] else None
        station_code = raw_row[header_index["stationCode"]] if "stationCode" in header_index and len(raw_row) > header_index["stationCode"] else default_station_code
        is_auto_filled = raw_row[header_index["isAutoFilled"]] if "isAutoFilled" in header_index and len(raw_row) > header_index["isAutoFilled"] else ""
        if phase in (None, "") and time_value in (None, "") and rainfall_value in (None, ""):
            continue
        row = {
            "phase": str(phase or "").strip().lower(),
            "time": time_value,
            "rainfallValue": rainfall_value,
            "stationCode": station_code,
            "isAutoFilled": str(is_auto_filled or "").strip(),
        }
        if row["phase"] == "history":
            history_rows.append(row)
        elif row["phase"] == "forecast":
            forecast_rows.append(row)
        else:
            raise ValueError(f"工作表 {sheet.title} 中存在未知 phase: {phase}")
    return history_rows, forecast_rows


def read_standard_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    if "metadata" not in workbook.sheetnames:
        raise ValueError("标准中间 Excel 缺少工作表: metadata")

    metadata = _read_key_value_sheet(workbook["metadata"])
    default_station_code = metadata.get("defaultStationCode", "").strip() or DEFAULT_STATION_CODE
    notes = []
    if "notes" in workbook.sheetnames:
        for row in workbook["notes"].iter_rows(min_row=2, values_only=True):
            note = "" if row[0] is None else str(row[0]).strip()
            if note:
                notes.append(note)

    if {"history_rainfall", "forecast_rainfall"}.issubset(workbook.sheetnames):
        return {
            "metadata": metadata,
            "history_rows": _read_legacy_rainfall_sheet(workbook["history_rainfall"]),
            "forecast_rows": _read_legacy_rainfall_sheet(workbook["forecast_rainfall"]),
            "notes": notes,
        }

    history_rows: list[dict[str, Any]] = []
    forecast_rows: list[dict[str, Any]] = []
    station_sheet_names = [sheet_name for sheet_name in workbook.sheetnames if sheet_name not in RESERVED_SHEETS]
    if not station_sheet_names:
        raise ValueError("标准中间 Excel 缺少 stationCode 对应工作表")

    for sheet_name in station_sheet_names:
        sheet_history_rows, sheet_forecast_rows = _read_station_sheet(workbook[sheet_name], default_station_code)
        history_rows.extend(sheet_history_rows)
        forecast_rows.extend(sheet_forecast_rows)
    return {
        "metadata": metadata,
        "history_rows": history_rows,
        "forecast_rows": forecast_rows,
        "notes": notes,
    }


def generate_time_series(
    *,
    forecast_time: str,
    history_duration: int,
    future_duration: int,
    time_step_hours: int,
    history_values: list[float],
    future_values: list[float],
    station_code: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    anchor = parse_datetime(forecast_time)
    step = timedelta(hours=time_step_hours)

    history_rows: list[dict[str, Any]] = []
    history_start = anchor - step * max(history_duration - 1, 0)
    for index, rainfall in enumerate(history_values):
        history_rows.append(
            {
                "time": format_datetime(history_start + step * index),
                "rainfallValue": float(rainfall),
                "stationCode": station_code,
                "isAutoFilled": "N",
            }
        )

    forecast_rows: list[dict[str, Any]] = []
    forecast_start = anchor + step
    for index, rainfall in enumerate(future_values):
        forecast_rows.append(
            {
                "time": format_datetime(forecast_start + step * index),
                "rainfallValue": float(rainfall),
                "stationCode": station_code,
                "isAutoFilled": "N",
            }
        )
    return history_rows, forecast_rows


def parse_json_array(raw: str | None, field_name: str) -> list[Any] | None:
    if raw is None:
        return None
    try:
        value = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{field_name} 不是合法 JSON 数组: {exc}") from exc
    if not isinstance(value, list):
        raise ValueError(f"{field_name} 必须是 JSON 数组")
    return value


def infer_forecast_time(history_rows: list[dict[str, Any]], forecast_rows: list[dict[str, Any]], time_step_hours: int) -> str:
    step = timedelta(hours=time_step_hours)
    if history_rows:
        anchor = max(parse_datetime(row["time"]) for row in history_rows)
        return format_datetime(anchor)
    if forecast_rows:
        anchor = min(parse_datetime(row["time"]) for row in forecast_rows) - step
        return format_datetime(anchor)
    raise ValueError("无法推断 forecastTime，历史和未来数据都为空")


def infer_durations(history_rows: list[dict[str, Any]], forecast_rows: list[dict[str, Any]], forecast_time: str, time_step_hours: int) -> tuple[int, int]:
    history_rows = align_rows_to_grid(history_rows, forecast_time=forecast_time, time_step_hours=time_step_hours)
    forecast_rows = align_rows_to_grid(forecast_rows, forecast_time=forecast_time, time_step_hours=time_step_hours)
    anchor = parse_datetime(forecast_time)
    step = timedelta(hours=time_step_hours)
    history_duration = 0
    future_duration = 0

    if history_rows:
        min_history_time = min(parse_datetime(row["time"]) for row in history_rows)
        history_duration = int((anchor - min_history_time) / step) + 1
    if forecast_rows:
        max_forecast_time = max(parse_datetime(row["time"]) for row in forecast_rows)
        future_duration = int((max_forecast_time - anchor) / step)

    if history_duration <= 0:
        raise ValueError("无法从输入推断 historyDuration，请确保历史数据至少包含 1 行")
    if future_duration <= 0:
        raise ValueError("无法从输入推断 futureDuration，请确保未来数据至少包含 1 行")
    return history_duration, future_duration


def expand_rows_by_station(
    *,
    history_rows: list[dict[str, Any]],
    forecast_rows: list[dict[str, Any]],
    forecast_time: str,
    history_duration: int,
    future_duration: int,
    time_step_hours: int,
    default_station_code: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    history_rows = align_rows_to_grid(history_rows, forecast_time=forecast_time, time_step_hours=time_step_hours)
    forecast_rows = align_rows_to_grid(forecast_rows, forecast_time=forecast_time, time_step_hours=time_step_hours)
    anchor = parse_datetime(forecast_time)
    step = timedelta(hours=time_step_hours)
    history_template = [format_datetime(anchor - step * (history_duration - 1 - index)) for index in range(history_duration)]
    forecast_template = [format_datetime(anchor + step * (index + 1)) for index in range(future_duration)]

    station_codes = set(group_rows_by_station(history_rows, default_station_code).keys()) | set(group_rows_by_station(forecast_rows, default_station_code).keys())
    if not station_codes:
        raise ValueError("没有可用的 stationCode 数据")

    notes: list[str] = []
    expanded_history: list[dict[str, Any]] = []
    expanded_forecast: list[dict[str, Any]] = []

    for station_code in sorted(station_codes):
        history_map: dict[str, dict[str, Any]] = {}
        forecast_map: dict[str, dict[str, Any]] = {}

        for row in history_rows:
            if normalize_station_code(row.get("stationCode"), default_station_code) != station_code:
                continue
            time_value = format_datetime(parse_datetime(row["time"]))
            if time_value in history_map:
                raise ValueError(f"stationCode={station_code} 的历史降雨存在重复时刻: {time_value}")
            history_map[time_value] = row

        for row in forecast_rows:
            if normalize_station_code(row.get("stationCode"), default_station_code) != station_code:
                continue
            time_value = format_datetime(parse_datetime(row["time"]))
            if time_value in forecast_map:
                raise ValueError(f"stationCode={station_code} 的未来降雨存在重复时刻: {time_value}")
            forecast_map[time_value] = row

        station_filled = 0
        for time_value in history_template:
            existing = history_map.get(time_value)
            if existing is None:
                station_filled += 1
                expanded_history.append({"time": time_value, "rainfallValue": 0.0, "stationCode": station_code, "isAutoFilled": "Y"})
            else:
                expanded_history.append({
                    "time": time_value,
                    "rainfallValue": coerce_rainfall_value(existing.get("rainfallValue")),
                    "stationCode": station_code,
                    "isAutoFilled": str(existing.get("isAutoFilled") or "N"),
                })

        for time_value in forecast_template:
            existing = forecast_map.get(time_value)
            if existing is None:
                station_filled += 1
                expanded_forecast.append({"time": time_value, "rainfallValue": 0.0, "stationCode": station_code, "isAutoFilled": "Y"})
            else:
                expanded_forecast.append({
                    "time": time_value,
                    "rainfallValue": coerce_rainfall_value(existing.get("rainfallValue")),
                    "stationCode": station_code,
                    "isAutoFilled": str(existing.get("isAutoFilled") or "N"),
                })

        if station_filled:
            notes.append(f"stationCode={station_code} 自动补齐 {station_filled} 个缺失时段，rainfallValue 设为 0。")

    return expanded_history, expanded_forecast, notes


def validate_station_counts(
    *,
    history_rows: list[dict[str, Any]],
    forecast_rows: list[dict[str, Any]],
    history_duration: int,
    future_duration: int,
    default_station_code: str,
) -> list[str]:
    errors: list[str] = []
    history_grouped = group_rows_by_station(history_rows, default_station_code)
    forecast_grouped = group_rows_by_station(forecast_rows, default_station_code)
    station_codes = sorted(set(history_grouped.keys()) | set(forecast_grouped.keys()))
    for station_code in station_codes:
        if len(history_grouped.get(station_code, [])) != history_duration:
            errors.append(
                f"stationCode={station_code} 的 history 行数为 {len(history_grouped.get(station_code, []))}，应为 {history_duration}"
            )
        if len(forecast_grouped.get(station_code, [])) != future_duration:
            errors.append(
                f"stationCode={station_code} 的 forecast 行数为 {len(forecast_grouped.get(station_code, []))}，应为 {future_duration}"
            )
    return errors
