#!/usr/bin/env python3
"""把靖州水文模型结构化 JSON 结果导出为 Excel 文件。"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


PREFERRED_DETAIL_COLUMNS = [
    "time",
    "flow",
    "modelType",
    "sectionName",
    "sectionCode",
    "stationCode",
    "rainfallValue",
]


def load_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise SystemExit(f"输入文件不存在: {path}") from exc
    except json.JSONDecodeError as exc:
        raise SystemExit(f"输入文件不是合法 JSON: {exc}") from exc

    if not isinstance(payload, dict):
        raise SystemExit("输入文件必须是 JSON 对象")
    return payload


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def collect_detail_columns(rows: list[dict[str, Any]]) -> list[str]:
    ordered = [column for column in PREFERRED_DETAIL_COLUMNS if any(column in row for row in rows)]
    extras: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in ordered and key not in extras:
                extras.append(key)
    return ordered + extras


def build_detail_rows(payload: dict[str, Any]) -> tuple[list[str], list[list[str]]]:
    response = payload.get("response") if isinstance(payload.get("response"), dict) else {}
    raw_rows = response.get("data")
    if not isinstance(raw_rows, list):
        return [], []

    detail_rows = [row for row in raw_rows if isinstance(row, dict)]
    if not detail_rows:
        return [], []

    columns = collect_detail_columns(detail_rows)
    rows = [[stringify(row.get(column)) for column in columns] for row in detail_rows]
    return columns, rows


def choose_section_name(row: dict[str, Any]) -> str:
    for key in ("sectionName", "section_name", "section", "sectionCode", "stationCode"):
        value = row.get(key)
        if value not in (None, ""):
            return str(value)
    return "未分组断面"


def sanitize_sheet_title(title: str) -> str:
    cleaned = "".join("_" if ch in '[]:*?/\\' else ch for ch in str(title)).strip()
    return (cleaned or "Sheet")[0:31]


def make_unique_sheet_title(base_title: str, existing_titles: set[str]) -> str:
    candidate = sanitize_sheet_title(base_title)
    if candidate not in existing_titles:
        existing_titles.add(candidate)
        return candidate

    counter = 2
    while True:
        suffix = f"_{counter}"
        trimmed = candidate[: max(0, 31 - len(suffix))].rstrip() or "Sheet"
        deduped = f"{trimmed}{suffix}"
        if deduped not in existing_titles:
            existing_titles.add(deduped)
            return deduped
        counter += 1


def group_rows_by_section(rows: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    ordered_names: list[str] = []
    for row in rows:
        section_name = choose_section_name(row)
        if section_name not in grouped:
            grouped[section_name] = []
            ordered_names.append(section_name)
        grouped[section_name].append(row)
    return [(name, grouped[name]) for name in ordered_names]


def write_table_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title=title)
    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=11)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap_alignment

    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    if headers:
        sheet.freeze_panes = "A2"


def export_to_excel(input_path: Path, output_path: Path) -> tuple[int, int]:
    payload = load_json(input_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    detail_headers, detail_rows = build_detail_rows(payload)
    if not detail_headers or not detail_rows:
        write_table_sheet(workbook, "结果明细", ["message"], [["响应中没有 data 明细"]])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return 0, 1

    raw_rows = [dict(zip(detail_headers, row)) for row in detail_rows]
    grouped_rows = group_rows_by_section(raw_rows)
    sheet_titles: set[str] = set()
    for section_name, section_rows in grouped_rows:
        sheet_title = make_unique_sheet_title(section_name, sheet_titles)
        rows = [[stringify(row.get(column)) for column in detail_headers] for row in section_rows]
        write_table_sheet(workbook, sheet_title, detail_headers, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(detail_rows), len(grouped_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="把靖州水文模型结构化 JSON 结果导出为 Excel 文件")
    parser.add_argument("--input_file", required=True, help="run_jingzhou_hydro_model.py 生成的结构化 JSON 文件")
    parser.add_argument("--output_file", help="输出 Excel 文件路径，默认与输入文件同名 .xlsx")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_file)
    output_path = Path(args.output_file) if args.output_file else input_path.with_suffix(".xlsx")
    detail_count, sheet_count = export_to_excel(input_path, output_path)
    print(f"Excel 文件已创建: {output_path}")
    print(f"结果明细行数: {detail_count}")
    print(f"生成工作表数量: {sheet_count}")


if __name__ == "__main__":
    main()
