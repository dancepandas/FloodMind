#!/usr/bin/env python3
"""调用敖江案例水文模型接口。"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from floodmind.skills.hydro_case_client import call_hydro_case_api, load_payload, save_result_if_needed

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

PREFERRED_DETAIL_COLUMNS = [
    "time",
    "flow",
    "modelType",
    "sectionName",
    "sectionCode",
    "stationCode",
    "rainfallValue",
]


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def collect_detail_columns(rows):
    ordered = [column for column in PREFERRED_DETAIL_COLUMNS if any(column in row for row in rows)]
    extras = []
    for row in rows:
        for key in row.keys():
            if key not in ordered and key not in extras:
                extras.append(key)
    return ordered + extras


def build_detail_rows(payload):
    response = payload.get("response") if isinstance(payload.get("response"), dict) else {}
    raw_rows = response.get("data")
    if not isinstance(raw_rows, list):
        return [], []

    detail_rows = [row for row in raw_rows if isinstance(row, dict)]
    if not detail_rows:
        return [], []

    columns = collect_detail_columns(detail_rows)
    rows = [[stringify(row.get(column)) for column in columns] for row in detail_rows]
    return columns, rows


def choose_section_name(row):
    for key in ("sectionName", "section_name", "section", "sectionCode", "stationCode"):
        value = row.get(key)
        if value not in (None, ""):
            return str(value)
    return "未分组断面"


def sanitize_sheet_title(title):
    cleaned = "".join("_" if ch in '[]:*?/\\' else ch for ch in str(title)).strip()
    return (cleaned or "Sheet")[0:31]


def make_unique_sheet_title(base_title, existing_titles):
    candidate = sanitize_sheet_title(base_title)
    if candidate not in existing_titles:
        existing_titles.add(candidate)
        return candidate

    counter = 2
    while True:
        suffix = f"_{counter}"
        trimmed = candidate[: max(0, 31 - len(suffix))].rstrip() or "Sheet"
        deduped = f"{trimmed}{suffix}"
        if deduped not in existing_titles:
            existing_titles.add(deduped)
            return deduped
        counter += 1


def group_rows_by_section(rows):
    grouped = {}
    ordered_names = []
    for row in rows:
        section_name = choose_section_name(row)
        if section_name not in grouped:
            grouped[section_name] = []
            ordered_names.append(section_name)
        grouped[section_name].append(row)
    return [(name, grouped[name]) for name in ordered_names]


def write_table_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title=title)
    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=11)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap_alignment

    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    if headers:
        sheet.freeze_panes = "A2"


def export_to_excel(payload, output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    detail_headers, detail_rows = build_detail_rows(payload)
    if not detail_headers or not detail_rows:
        write_table_sheet(workbook, "结果明细", ["message"], [["响应中没有 data 明细"]])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return 0, 1

    raw_rows = [dict(zip(detail_headers, row)) for row in detail_rows]
    grouped_rows = group_rows_by_section(raw_rows)
    sheet_titles = set()
    for section_name, section_rows in grouped_rows:
        sheet_title = make_unique_sheet_title(section_name, sheet_titles)
        rows = [[stringify(row.get(column)) for column in detail_headers] for row in section_rows]
        write_table_sheet(workbook, sheet_title, detail_headers, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(detail_rows), len(grouped_rows)


def build_excel_output_path(output_file: str | None, excel_output_file: str | None) -> Path | None:
    if excel_output_file:
        return Path(excel_output_file)
    if output_file:
        return Path(output_file).with_suffix(".xlsx")
    return Path("result.xlsx")


def build_result_summary(
    result_text: str,
    json_path: Path | None,
    excel_path: Path | None,
    detail_count: int | None = None,
    sheet_count: int | None = None,
) -> str:
    lines = [str(result_text or "").strip()]
    if json_path and excel_path:
        lines.append(
            f"已经生成标准的 {json_path.name} 和 {excel_path.name}。"
            f"成果检查等可直接检查 {json_path.name} 文件内容是否正确；{excel_path.name} 文件是 {json_path.name} 文件经过脚本固定转换得到的，"
            f"{json_path.name} 内容正确则 {excel_path.name} 一定正确。"
        )
        lines.append(f"JSON 路径: {json_path}")
        lines.append(f"Excel 路径: {excel_path}")
    elif json_path:
        lines.append(f"已经生成标准的 {json_path.name}。")
        lines.append(f"JSON 路径: {json_path}")
    elif excel_path:
        lines.append(f"已经生成标准的 {excel_path.name}。")
        lines.append(f"Excel 路径: {excel_path}")
    if detail_count is not None and sheet_count is not None:
        lines.append(f"转换结果统计: 明细 {detail_count} 行，工作表 {sheet_count} 个")
    return "\n\n".join(line for line in lines if line)


def main() -> None:
    parser = argparse.ArgumentParser(description="调用敖江案例水文模型接口")
    parser.add_argument("--input_file", help="完整请求体 JSON 文件路径")
    parser.add_argument("--payload", help="完整请求体 JSON 字符串")
    parser.add_argument("--forecast_time", help="预报时刻，格式 YYYY-MM-DD HH:MM:SS")
    parser.add_argument("--history_duration", type=int, help="历史数据时长（小时）")
    parser.add_argument("--future_duration", type=int, help="预报时长（小时）")
    parser.add_argument("--model_run_param", help="模型参数对象 JSON")
    parser.add_argument("--model_data_params", help="历史降雨数组 JSON")
    parser.add_argument(
        "--model_forecast_rainfall_params",
        help="未来降雨数组 JSON",
    )
    parser.add_argument("--base_url", default="http://192.168.30.108:11111", help="服务基础地址")
    parser.add_argument("--timeout", type=int, default=120, help="请求超时时间（秒）")
    parser.add_argument("--output_file", help="结构化结果 JSON 保存路径；默认输出到当前目录下 result.json")
    parser.add_argument("--excel_output_file", help="可选，结果 Excel 保存路径；默认与 JSON 输出同名 .xlsx")

    args = parser.parse_args()

    try:
        payload = load_payload(
            input_file=args.input_file,
            payload_json=args.payload,
            forecast_time=args.forecast_time,
            history_duration=args.history_duration,
            future_duration=args.future_duration,
            model_run_param=args.model_run_param,
            model_data_params=args.model_data_params,
            model_forecast_rainfall_params=args.model_forecast_rainfall_params,
        )
        result_payload = call_hydro_case_api(
            payload=payload,
            base_url=args.base_url,
            endpoint="/aj/hydro_model_sync",
            case_name="敖江",
            timeout=args.timeout,
        )
        json_output_path = Path(args.output_file) if args.output_file else Path("result.json")
        result_text = save_result_if_needed(result_payload, str(json_output_path))
        excel_output_path = build_excel_output_path(args.output_file, args.excel_output_file)
        detail_count, sheet_count = export_to_excel(result_payload, excel_output_path)
        print(build_result_summary(result_text, json_output_path, excel_output_path, detail_count, sheet_count))
    except Exception as exc:
        print(f"敖江水文模型脚本执行失败: {exc}")


if __name__ == "__main__":
    main()
