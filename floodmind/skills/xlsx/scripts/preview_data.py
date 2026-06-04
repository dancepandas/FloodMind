"""
数据预览脚本 - 仅读取文件前N行用于预览

安全措施：
1. 默认只读取5行
2. 最大限制100行（防止内存溢出）
3. 文件大小限制50MB
4. 不加载完整数据到内存
"""

import argparse
import os
import sys
from pathlib import Path


def get_file_sizeKB(path: str) -> float:
    return os.path.getsize(path) / 1024


def preview_csv(file_path: str, n_rows: int = 5, encoding: str = "utf-8") -> dict:
    import csv

    rows = []
    total_rows = 0

    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            total_rows += 1
            if i < n_rows:
                rows.append(row)
            elif i >= 10000:
                break

    return {
        "total_rows": total_rows if total_rows <= 10000 else f">10000 (文件过大，仅统计前10000行)",
        "preview_rows": len(rows),
        "columns": len(rows[0]) if rows else 0,
        "headers": rows[0] if rows else [],
        "data": rows[1:] if len(rows) > 1 else [],
    }


def preview_excel(file_path: str, n_rows: int = 5, sheet_name: str = None) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return {"error": f"工作表 '{sheet_name}' 不存在，可用: {wb.sheetnames}"}
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = []
    total_rows = 0
    max_rows = 10000

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        total_rows += 1
        if i < n_rows:
            rows.append([str(c) if c is not None else "" for c in row])
        elif i >= max_rows:
            break

    wb.close()

    return {
        "total_rows": total_rows if total_rows <= max_rows else f">{max_rows}",
        "preview_rows": len(rows),
        "columns": len(rows[0]) if rows else 0,
        "headers": rows[0] if rows else [],
        "data": rows[1:] if len(rows) > 1 else [],
        "sheet_names": wb.sheetnames if not sheet_name else None,
    }


def preview_json(file_path: str, n_rows: int = 5, encoding: str = "utf-8") -> dict:
    import json

    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        content = f.read(1024 * 1024)

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        return {"error": "JSON解析失败，文件可能不是标准JSON格式"}

    if isinstance(data, list):
        rows = data[:n_rows]
        return {
            "total_rows": len(data) if len(data) <= 1000 else f">{len(data)} (文件过大)",
            "preview_rows": len(rows),
            "columns": list(data[0].keys()) if data and isinstance(data[0], dict) else [],
            "data": rows,
        }
    else:
        return {
            "total_rows": 1,
            "preview_rows": 1,
            "columns": list(data.keys()) if isinstance(data, dict) else [],
            "data": [data],
        }


def preview_txt(file_path: str, n_rows: int = 5, encoding: str = "utf-8") -> dict:
    lines = []
    total_lines = 0

    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        for i, line in enumerate(f):
            total_lines += 1
            if i < n_rows:
                lines.append(line.rstrip("\n\r"))
            elif i >= 10000:
                break

    return {
        "total_rows": total_lines if total_lines <= 10000 else f">10000 (文件过大)",
        "preview_rows": len(lines),
        "columns": 1,
        "headers": ["Text"],
        "data": [[l] for l in lines],
    }


def main():
    parser = argparse.ArgumentParser(description="数据预览 - 仅读取前N行")
    parser.add_argument("--file_path", required=True, help="文件路径")
    parser.add_argument("--n_rows", type=int, default=5, help="预览行数，默认5，最大100")
    parser.add_argument("--sheet_name", default=None, help="Excel工作表名称")
    parser.add_argument("--encoding", default="utf-8", help="文件编码")

    args = parser.parse_args()

    file_path = Path(args.file_path)

    if not file_path.exists():
        print(f"错误：文件不存在: {file_path}")
        sys.exit(1)

    size_kb = get_file_sizeKB(file_path)
    if size_kb > 51200:
        print(f"错误：文件过大 ({size_kb:.1f}KB)，最大支持 50MB")
        sys.exit(1)

    n_rows = min(max(1, args.n_rows), 100)

    ext = file_path.suffix.lower()

    try:
        if ext == ".csv":
            result = preview_csv(str(file_path), n_rows, args.encoding)
        elif ext in [".xlsx", ".xls", ".xlsm"]:
            result = preview_excel(str(file_path), n_rows, args.sheet_name)
        elif ext == ".json":
            result = preview_json(str(file_path), n_rows, args.encoding)
        elif ext == ".txt":
            result = preview_txt(str(file_path), n_rows, args.encoding)
        else:
            result = preview_txt(str(file_path), n_rows, args.encoding)
    except Exception as e:
        print(f"错误：读取文件失败: {e}")
        sys.exit(1)

    if "error" in result:
        print(f"错误：{result['error']}")
        sys.exit(1)

    print("=" * 60)
    print(f"文件: {file_path.name}")
    print(f"总行数: {result['total_rows']}")
    print(f"预览行数: {result['preview_rows']}")
    print(f"列数: {result['columns']}")
    print("=" * 60)

    print(f"\n【表头】")
    print(" | ".join(str(h) for h in result["headers"]))

    print(f"\n【数据预览（前{result['preview_rows']-1}行）】")
    for row in result["data"]:
        print(" | ".join(str(c)[:30] for c in row))

    if result.get("sheet_names"):
        print(f"\n【Excel工作表】{result['sheet_names']}")

    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
