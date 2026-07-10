"""
文件工具

上传 / 输出目录管理、文件名安全化、产物提取、预览构建。
"""
import hashlib
import json
import logging
import mimetypes
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

from floodmind.server.config import (
    ARTIFACT_EXTENSIONS, ARTIFACT_FILENAME_PATTERN, ARTIFACT_PATH_PATTERN,
    DOWNLOADABLE_EXTENSIONS, IMAGE_EXTENSIONS,
)

logger = logging.getLogger(__name__)

# ── 会话级文件缓存 ────────────────────────────────────
import threading
session_files: Dict[str, Dict[str, dict]] = {}
session_files_lock = threading.RLock()


def _require_session_id(raw: Optional[str]) -> str:
    from floodmind.memory.session_manager import validate_session_id
    return validate_session_id(raw or "default")


def _is_within_dir(path: str, base_dir: str) -> bool:
    try:
        return os.path.commonpath([os.path.realpath(path), os.path.realpath(base_dir)]) == os.path.realpath(base_dir)
    except ValueError:
        return False


def _stable_upload_file_id(filename: str) -> str:
    return hashlib.sha1(str(filename).encode("utf-8")).hexdigest()[:16]


def _get_upload_index_path(session_id: str, session_manager: "SessionManager") -> Path:
    return session_manager.get_session_dir(session_id) / "uploads_index.json"


def get_session_upload_dir(session_id: str, session_manager: "SessionManager") -> str:
    session_id = _require_session_id(session_id)
    return str(session_manager.get_upload_dir(session_id))


def get_session_output_dir(session_id: str, session_manager: "SessionManager") -> str:
    session_id = _require_session_id(session_id)
    return str(session_manager.get_output_dir(session_id))


# ── 文件名处理 ─────────────────────────────────────────

def allowed_file(filename: str) -> bool:
    """检查文件扩展名是否允许。"""
    from floodmind.server.config import ALLOWED_EXTENSIONS
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def safe_filename(filename: str) -> str:
    """保留中文的文件名安全化处理，仅移除路径分隔符等危险字符。"""
    name, ext = os.path.splitext(filename)
    name = re.sub(r'[\\/:*?"<>|]', '_', name).strip()
    if not name:
        name = 'file'
    return f"{name}{ext}"


def dedup_filename(directory: str, filename: str) -> str:
    """若目录下已存在同名文件，自动加编号，如 文件(1).docx。"""
    target = os.path.join(directory, filename)
    if not os.path.exists(target):
        return filename
    name, ext = os.path.splitext(filename)
    counter = 1
    while os.path.exists(os.path.join(directory, f"{name}({counter}){ext}")):
        counter += 1
    return f"{name}({counter}){ext}"


# ── 上传文件索引 ───────────────────────────────────────

def _save_session_files(session_id: str, session_manager: "SessionManager") -> None:
    with session_files_lock:
        records = list(session_files.get(session_id, {}).values())
    _get_upload_index_path(session_id, session_manager).write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_session_files(session_id: str, session_manager: "SessionManager") -> Dict[str, dict]:
    session_id = _require_session_id(session_id)
    uploads_dir = Path(get_session_upload_dir(session_id, session_manager))
    index_path = _get_upload_index_path(session_id, session_manager)
    records: Dict[str, dict] = {}

    if index_path.exists():
        try:
            data = json.loads(index_path.read_text(encoding="utf-8"))
            if isinstance(data, list):
                for item in data:
                    if not isinstance(item, dict):
                        continue
                    file_id = str(item.get("id", "")).strip()
                    file_name = str(item.get("name", "")).strip()
                    file_path = str(item.get("path", "")).strip()
                    if not file_id or not file_name or not file_path or not os.path.isfile(file_path):
                        continue
                    records[file_id] = {
                        "id": file_id, "name": file_name, "path": file_path,
                        "size": int(item.get("size", 0) or 0),
                        "upload_time": str(item.get("upload_time", "") or ""),
                    }
        except Exception as e:
            logger.warning("读取上传文件索引失败: %s", e)

    if not records and uploads_dir.exists():
        for path in sorted(uploads_dir.iterdir()):
            if not path.is_file():
                continue
            file_id = _stable_upload_file_id(path.name)
            records[file_id] = {
                "id": file_id, "name": path.name, "path": str(path),
                "size": path.stat().st_size,
                "upload_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
            }

    with session_files_lock:
        session_files[session_id] = records
    if records:
        _save_session_files(session_id, session_manager)
    return records


def get_session_files_map(session_id: str, session_manager: "SessionManager") -> Dict[str, dict]:
    session_id = _require_session_id(session_id)
    with session_files_lock:
        cached = session_files.get(session_id)
    if cached is not None:
        return cached
    return _load_session_files(session_id, session_manager)


def remove_session_files(session_id: str, session_manager: "SessionManager") -> None:
    with session_files_lock:
        session_files.pop(session_id, None)
    index_path = _get_upload_index_path(session_id, session_manager)
    if index_path.exists():
        index_path.unlink()


def public_file_info(file_info: dict) -> dict:
    return {
        'id': file_info.get('id', ''),
        'name': file_info.get('name', ''),
        'size': file_info.get('size', 0),
        'upload_time': file_info.get('upload_time', ''),
    }


# ── 产物 URL / 预览 ────────────────────────────────────

def build_session_output_url(filepath: str, fallback_session_id: str, session_manager: "SessionManager") -> str:
    """根据输出文件路径构造会话输出 URL，保留子目录相对路径。"""
    real_path = os.path.realpath(filepath)
    output_dir = os.path.realpath(get_session_output_dir(fallback_session_id, session_manager))
    if _is_within_dir(real_path, output_dir):
        rel = os.path.relpath(real_path, output_dir).replace(os.sep, "/")
        return f"/api/sessions/{fallback_session_id}/outputs/{rel}"
    return f"/api/sessions/{fallback_session_id}/outputs/{os.path.basename(filepath)}"


def build_uploaded_file_preview(file_info: dict) -> dict:
    """安全读取上传文件预览内容，不暴露 uploads 原始路径。"""
    file_path = file_info.get('path', '')
    file_name = file_info.get('name', '')
    ext = os.path.splitext(file_name)[1].lower()

    preview = {
        'file_id': file_info.get('id', ''),
        'file_name': file_name,
        'size': file_info.get('size', 0),
        'preview_type': 'unsupported',
        'content': '',
    }

    if not os.path.exists(file_path):
        preview['preview_type'] = 'missing'
        preview['content'] = '文件不存在或已被清理。'
        return preview

    if ext in {'.txt', '.json'}:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as handle:
            preview['content'] = handle.read(6000)
        preview['preview_type'] = 'text'
        return preview

    if ext == '.csv':
        import csv
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                rows_data = []
                for i, row in enumerate(reader):
                    if i >= 20:
                        break
                    rows_data.append(row)
            if rows_data:
                preview['preview_type'] = 'table'
                preview['columns'] = rows_data[0] if rows_data else []
                preview['rows'] = rows_data[1:] if len(rows_data) > 1 else []
            else:
                preview['preview_type'] = 'text'
                preview['content'] = '(空文件)'
        except Exception as e:
            preview['preview_type'] = 'text'
            preview['content'] = f'CSV 预览失败: {e}'
        return preview

    if ext in {'.xlsx', '.xls'}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            preview['preview_type'] = 'text'
            preview['content'] = 'Excel 预览需要 openpyxl 库。'
            return preview
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            preview['preview_type'] = 'excel'
            preview['sheets'] = []
            for sheet_name in wb.sheetnames[:5]:
                ws = wb[sheet_name]
                rows_data = []
                for i, row in enumerate(ws.iter_rows(values_only=True, max_row=12)):
                    rows_data.append([str(c) if c is not None else '' for c in row])
                preview['sheets'].append({
                    'sheet_name': sheet_name,
                    'columns': rows_data[0] if rows_data else [],
                    'rows': rows_data[1:] if len(rows_data) > 1 else [],
                })
            wb.close()
        except Exception as e:
            preview['preview_type'] = 'text'
            preview['content'] = f'Excel 预览失败: {e}'
        return preview

    if ext == '.md':
        with open(file_path, 'r', encoding='utf-8', errors='replace') as handle:
            preview['content'] = handle.read(6000)
        preview['preview_type'] = 'text'
        return preview

    if ext in {'.docx', '.pdf'}:
        preview['preview_type'] = 'document'
        preview['content'] = '点击预览按钮查看文档内容。'
        return preview

    preview['content'] = '该文件类型暂不支持在线预览。'
    return preview


# ── 产物提取 ───────────────────────────────────────────

def extract_generated_paths(content: str) -> List[str]:
    """从工具输出中提取生成文件路径。"""
    if not content:
        return []
    return [match.strip().strip('*').strip('-').strip('`').strip()
            for match in ARTIFACT_PATH_PATTERN.findall(content)]


def extract_generated_filenames(content: str) -> List[str]:
    """从文本中提取仅包含文件名的成果引用。"""
    if not content:
        return []
    seen: List[str] = []
    for match in ARTIFACT_FILENAME_PATTERN.findall(content):
        filename = str(match).strip().strip('`').strip()
        if filename and filename not in seen:
            seen.append(filename)
    return seen


def build_artifact_event(
    filepath: str, fallback_session_id: str, emitted_paths: Set[str],
    session_manager: "SessionManager",
) -> Optional[dict]:
    """为新生成的输出文件构造 SSE 事件。"""
    real_path = os.path.realpath(filepath)
    ext = os.path.splitext(real_path)[1].lower()

    if ext not in ARTIFACT_EXTENSIONS:
        return None
    if not os.path.exists(real_path):
        logger.warning("生成文件不存在: %s", real_path)
        return None

    output_dir = os.path.realpath(get_session_output_dir(fallback_session_id, session_manager))
    if not _is_within_dir(real_path, output_dir):
        logger.warning("生成文件不在会话输出目录内: %s", real_path)
        return None
    if real_path in emitted_paths:
        return None

    emitted_paths.add(real_path)
    url = build_session_output_url(real_path, fallback_session_id, session_manager)
    file_info: Dict[str, Any] = {
        'filename': os.path.basename(real_path),
        'size': os.path.getsize(real_path),
    }

    if ext in IMAGE_EXTENSIONS:
        file_info['type'] = 'image_generated'
        file_info['image_url'] = url
        file_info['download_url'] = url
    else:
        file_info['type'] = 'file_generated'
        file_info['download_url'] = url

    return file_info


def list_recent_output_artifacts(output_dir: str, request_started_at: float) -> List[str]:
    """列出本轮请求期间新生成的输出文件。"""
    if not os.path.isdir(output_dir):
        return []
    artifacts = []
    for filename in os.listdir(output_dir):
        full_path = os.path.join(output_dir, filename)
        if not os.path.isfile(full_path):
            continue
        ext = os.path.splitext(filename)[1].lower()
        if ext not in ARTIFACT_EXTENSIONS:
            continue
        try:
            if os.path.getmtime(full_path) >= request_started_at - 1:
                artifacts.append(full_path)
        except OSError as exc:
            logger.warning("读取输出文件时间失败: %s, %s", full_path, exc)
    artifacts.sort(key=lambda path: os.path.getmtime(path))
    return artifacts


def extract_validated_artifact_paths(content: str, session_id: str, session_manager: "SessionManager") -> List[str]:
    """只提取已经被调度器/校验逻辑认定为最终可交付的成果文件。"""
    text = (content or '').strip()
    if not text:
        return []
    try:
        payload = json.loads(text)
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []

    is_native_delegate = payload.get('is_native_delegate', False)
    validation = payload.get('validation') or {}
    overall_status = str(validation.get('overall_status', '')).strip().lower()
    is_final_goal_met = validation.get('is_final_goal_met')
    current_result_type = str(validation.get('current_result_type', '')).strip().lower()
    if not is_native_delegate and overall_status != 'pass' and is_final_goal_met is not True:
        return []
    if not is_native_delegate and current_result_type and current_result_type != 'final_deliverable':
        return []

    artifacts = payload.get('artifacts') or []
    if not isinstance(artifacts, list):
        return []

    output_dir = get_session_output_dir(session_id, session_manager) if session_id else ""
    deliverable_exts = {'.xlsx', '.xls', '.docx', '.doc', '.pdf', '.png', '.jpg', '.jpeg', '.gif'}
    result = []
    for path in artifacts:
        normalized = str(path).strip()
        if not normalized:
            continue
        ext = os.path.splitext(normalized)[1].lower()
        if ext not in deliverable_exts:
            continue
        if output_dir and not _is_within_dir(os.path.realpath(normalized), os.path.realpath(output_dir)):
            continue
        result.append(normalized)
    return result


def filter_final_delivery_artifacts(artifact_paths: List[str], final_text: str) -> List[str]:
    """只保留最终回答里明确点名交付的成果文件。"""
    if not artifact_paths:
        return []
    normalized_text = (final_text or '').strip()
    if not normalized_text:
        return artifact_paths
    mentioned = []
    for path in artifact_paths:
        filename = os.path.basename(str(path))
        if filename and filename in normalized_text:
            mentioned.append(path)
    return mentioned or artifact_paths


def resolve_artifact_references(
    session_id: str, artifact_paths: List[str], final_text: str,
    session_manager: "SessionManager",
) -> List[str]:
    """综合完整路径和仅文件名引用，解析最终可交付文件。"""
    resolved: List[str] = []
    seen: Set[str] = set()
    for path in filter_final_delivery_artifacts(artifact_paths, final_text):
        normalized = os.path.realpath(str(path))
        if normalized not in seen:
            seen.add(normalized)
            resolved.append(normalized)
    output_dir = get_session_output_dir(session_id, session_manager)
    for filename in extract_generated_filenames(final_text):
        candidate = os.path.realpath(os.path.join(output_dir, filename))
        if os.path.isfile(candidate) and candidate not in seen:
            seen.add(candidate)
            resolved.append(candidate)
    return resolved


# ── 产物持久化 ─────────────────────────────────────────

def _artifact_dedup_key(event: Dict[str, Any]) -> str:
    url = event.get('download_url') or event.get('image_url', '')
    if url:
        return str(url)
    return f"{event.get('type', '')}:{event.get('filename', '')}"


def list_session_artifact_events(session_id: str, session_manager: "SessionManager") -> List[Dict[str, Any]]:
    """列出某个会话最终确认可交付的产物事件。"""
    from floodmind.server.sanitize import sanitize_artifact_event
    artifacts_file = os.path.join(
        str(session_manager.get_session_dir(session_id)), 'approved_artifacts.json')
    if os.path.exists(artifacts_file):
        try:
            data = json.loads(Path(artifacts_file).read_text(encoding='utf-8'))
            if isinstance(data, list):
                items = [sanitize_artifact_event(item) for item in data if isinstance(item, dict)]
                for item in items:
                    filename = item.get('filename', '')
                    if not item.get('download_url') and filename:
                        item['download_url'] = f"/api/sessions/{session_id}/outputs/{filename}"
                    if item.get('type') == 'image_generated' and not item.get('image_url') and item.get('download_url'):
                        item['image_url'] = item['download_url']
                return items
        except Exception as e:
            logger.warning("读取最终成果文件记录失败: %s", e)
    return []


def save_session_artifact_events(
    session_id: str, artifact_events: List[Dict[str, Any]],
    session_manager: "SessionManager",
) -> None:
    """保存会话产物事件（去重合并）。"""
    from floodmind.server.sanitize import sanitize_artifact_event
    artifacts_file = os.path.join(
        str(session_manager.get_session_dir(session_id)), 'approved_artifacts.json')
    existing: List[Dict[str, Any]] = []
    if os.path.exists(artifacts_file):
        try:
            data = json.loads(Path(artifacts_file).read_text(encoding='utf-8'))
            if isinstance(data, list):
                existing = [item for item in data if isinstance(item, dict)]
        except Exception:
            pass
    new_sanitized = [sanitize_artifact_event(item) for item in artifact_events]
    seen: Set[str] = set()
    merged: List[Dict[str, Any]] = []
    for item in existing:
        key = _artifact_dedup_key(item)
        if key not in seen:
            seen.add(key)
            merged.append(item)
    for item in new_sanitized:
        key = _artifact_dedup_key(item)
        if key not in seen:
            seen.add(key)
            merged.append(item)
    Path(artifacts_file).write_text(
        json.dumps(merged, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )


def build_artifact_summary_text(
    artifact_events: List[Dict[str, Any]], original_text: str,
) -> Optional[str]:
    """基于本轮真实生成文件构造更面向用户的最终交付文案。"""
    if not artifact_events:
        return None
    lines = []
    image_events = [e for e in artifact_events if e.get('type') == 'image_generated']
    deliverable_exts = {'.xlsx', '.xls', '.docx', '.doc', '.pdf'}
    file_events = [
        e for e in artifact_events
        if e.get('type') == 'file_generated'
        and os.path.splitext(str(e.get('filename', '')))[1].lower() in deliverable_exts
    ]
    cleaned_original = (original_text or "").strip()
    if not cleaned_original:
        cleaned_original = "任务已完成。"
    internal_artifact_names = {'input.json', 'result.json', 'result.xlsx', 'result.csv', 'result.txt'}
    is_explicit_delivery = any(
        kw in cleaned_original for kw in (
            '最终交付', '已生成', '导出文件', '生成报告', '生成文件', '保存文件', '下载文件'
        )
    )
    if image_events:
        label = "最终交付图片" if is_explicit_delivery else "附带图片"
        lines.append(f"{label}：")
        for event in image_events:
            lines.append(f"- `{event.get('filename', '')}`")
    if file_events:
        internal_files = [e for e in file_events if e.get('filename', '') in internal_artifact_names]
        deliverable_files = [e for e in file_events if e not in internal_files]
        if deliverable_files:
            label = "最终交付文件" if is_explicit_delivery else "附带结果文件"
            lines.append(f"{label}：")
            for event in deliverable_files:
                lines.append(f"- `{event.get('filename', '')}`")
        if internal_files and not deliverable_files:
            lines.append("附带结果文件：")
            for event in internal_files:
                lines.append(f"- `{event.get('filename', '')}`")
    if not lines:
        return None
    summary_block = "\n".join(lines)
    if summary_block in cleaned_original:
        return cleaned_original
    return f"{cleaned_original}\n\n{summary_block}"
